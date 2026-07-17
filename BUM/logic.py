"""Core builder for the BUM tool.

Inputs:
  * Main AR dump  - as-of date in A1, headers on row 2, data from row 3.
  * AUH customers list   - report; header row starts with 'Cust Code',
    the 'Addr State Code' column carries the AUH flag.
  * Renewals invoices    - report; header row starts with 'Invoice number'.
  * Insurance master     - report; header row starts with 'Customer Code',
    the 'Insurance Limit' column carries the limit.
  * PDC due to be banked - report; header row starts with 'Division';
    pivoted here as Sub Account -> sum of LC Amount.
  * Sales backlog        - report; header row starts with 'Order';
    pivoted here as Customer Code -> sum of Pending Val (Lc).

Output workbook:
  * Sheet1            - the main dump plus the formula columns.
  * Sheet3            - raw copy of the AUH customers list.
  * Renewal invoices  - raw copy of the renewals report, plus a helper
    column G filled with "Renewals" (the VLOOKUP returns column 7).
  * Insurance         - raw copy of the insurance master report.
  * PDC Pivot / Backlog Pivot - pivots computed from the two reports
    (Excel-pivot layout: Row Labels from row 4, then (blank), Grand Total).
  * BUM fixed / Region / GSI / SE Africa - fixed lists bundled in BUM/data.

All added columns are live Excel formulas so the numbers can be audited in
the file itself.
"""

from __future__ import annotations

import calendar
import csv
import io
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import xlsxwriter
from openpyxl.utils import get_column_letter

_DATA_DIR = Path(__file__).parent / "data"

# Input headers the formulas depend on.
_REQUIRED_HEADERS = (
    "Cust Code",
    "Cust Region",
    "Document Number",
    "Over Due Days",
    "Document Due Date",
    "On Account",
    "Not Due Amount",
    "Ar Balance",
    "Brand",
)

NEW_HEADERS = [
    "Invoice Age",
    "Invoice Value",
    "3. Aging 1 to 15",
    "4. Aging 16 to 30",
    "5. Aging 31 to 60",
    "6. Aging 61 to 90",
    "7. Aging 91 to 120",
    "8. Aging >=121",
    "Additional due End of month",
    "9. AR Balance",
    "1. On Account",
    "2. Not Due",
    "AUH",
    "BUM",
    "Region",
    "Shelly",
    "Renewals",
    "SE Africa",
    "Insurance",
    "Backlog",
    "PDC",
]

_FMT_DATE = "mm-dd-yy"
_FMT_USD = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
_FMT_NUM = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
_FMT_NUM2 = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
_FMT_PLAIN = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'


def _read_csv(name: str) -> list[list[str]]:
    with open(_DATA_DIR / name, encoding="utf-8", newline="") as f:
        return [row for row in csv.reader(f) if row]


# Rows dropped from the main AR dump: internal/intercompany customers and
# excluded main accounts. Matching is case-insensitive; names match if the
# customer name CONTAINS the keyword.
EXCLUDED_CUST_NAME_KEYWORDS = ("MINDWARE", "AKLANIAT", "IFIX")
EXCLUDED_MAIN_AC = ("12302", "12304", "12306")


def _read_main(file_bytes: bytes):
    """Return (as_of, headers, data_rows) from the main AR dump.

    Tolerates banner rows: the header row is found by its first cell
    ('Cust Code') and the as-of date is read from the 'As on Date' banner
    row (falling back to cell A1 for the old layout). Rows for excluded
    customers / main accounts and the trailing summary row are dropped.
    """
    def _norm(v):
        return " ".join(str(v).replace("\xa0", " ").split()).lower()

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True
    )
    raw = None
    header_idx = col_off = None  # 0-based
    for ws in wb.worksheets:
        ws.reset_dimensions()  # some reports declare a bogus sheet size
        sheet_raw = [list(row) for row in ws.iter_rows(values_only=True)]
        for i, row in enumerate(sheet_raw[:60]):
            for j, cell in enumerate(row or ()):
                if isinstance(cell, str) and _norm(cell) == "cust code":
                    header_idx, col_off, raw = i, j, sheet_raw
                    break
            if raw is not None:
                break
        if raw is not None:
            break
    wb.close()
    if raw is None:
        raise ValueError(
            "Main file: could not find a 'Cust Code' header cell in the "
            "first 60 rows of any sheet."
        )

    header_cells = list(raw[header_idx][col_off:])
    while header_cells and header_cells[-1] is None:
        header_cells.pop()
    # Clean hidden spaces and snap known columns to their canonical spelling
    # so lookups by header name survive report formatting quirks.
    canonical = _REQUIRED_HEADERS + (
        "Cust Name", "Main Ac", "SO No", "LPO No", "Document Date",
        "Days From Docdt", "LC & BG Guarantee", "Customer Status",
    )
    canon_map = {_norm(h): h for h in canonical}
    headers = []
    for k, h in enumerate(header_cells):
        if h is None:
            headers.append(f"Column{k + 1}")
        else:
            cleaned = " ".join(str(h).replace("\xa0", " ").split())
            headers.append(canon_map.get(_norm(h), cleaned))

    missing = [h for h in _REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError(f"Main file is missing expected columns: {missing}")

    def _to_date(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, str) and v.strip():
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    return datetime.strptime(v.strip(), fmt).date()
                except ValueError:
                    pass
        return None

    as_of = None
    for row in raw[:header_idx]:
        for j, cell in enumerate(row or ()):
            if isinstance(cell, str) and _norm(cell).startswith("as on date"):
                for v in row[j + 1 :]:
                    as_of = _to_date(v)
                    if as_of:
                        break
                if as_of is None:  # value glued into the same cell
                    as_of = _to_date(cell.strip()[10:].strip(": "))
                break
        if as_of is not None:
            break
    if as_of is None:  # fallback: 'Run Date:17/07/2026 12:00' banner cell
        for row in raw[:header_idx]:
            for cell in row or ():
                if isinstance(cell, str) and _norm(cell).startswith("run date"):
                    as_of = _to_date(cell.split(":", 1)[-1].strip()[:10])
                    break
            if as_of is not None:
                break
    if as_of is None and raw and raw[0]:  # old layout: date in A1
        as_of = _to_date(raw[0][0])
    if as_of is None:
        raise ValueError(
            "Main file: could not find the as-of date (looked for an "
            "'As on Date' row, then 'Run Date', then cell A1)."
        )

    n_cols = len(headers)
    code_i = headers.index("Cust Code")
    name_i = headers.index("Cust Name")
    mac_i = headers.index("Main Ac")
    data = []
    for row in raw[header_idx + 1 :]:
        row = list(row[col_off : col_off + n_cols])
        row += [None] * (n_cols - len(row))
        if all(v is None or v == "" for v in row):
            continue
        code = row[code_i] if len(row) > code_i else None
        if code in (None, ""):  # summary/total rows carry no customer code
            continue
        if isinstance(code, str) and code.strip().lower().startswith("summary"):
            continue
        name = str(row[name_i] or "").upper() if len(row) > name_i else ""
        if any(k in name for k in EXCLUDED_CUST_NAME_KEYWORDS):
            continue
        mac = str(row[mac_i]).strip() if len(row) > mac_i and row[mac_i] is not None else ""
        if mac.endswith(".0"):
            mac = mac[:-2]
        if mac in EXCLUDED_MAIN_AC:
            continue
        data.append(row)
    return as_of, headers, data


class _Report:
    """A raw report whose header row is found by its first cell text."""

    def __init__(self, file_bytes: bytes, first_header: str, label: str):
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws = wb.worksheets[0]
        ws.reset_dimensions()  # some reports declare a bogus sheet size
        self.label = label
        self.raw = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

        header_idx = None  # 0-based
        for i, row in enumerate(self.raw[:30]):
            if row and isinstance(row[0], str) and row[0].strip() == first_header:
                header_idx = i
                break
        if header_idx is None:
            raise ValueError(
                f"{label} file: could not find the header row starting "
                f"with '{first_header}'."
            )
        self.headers = [
            str(h).strip() if h is not None else "" for h in self.raw[header_idx]
        ]
        # reset_dimensions() yields ragged rows (each only as long as its
        # last filled cell) - pad to header width so positional access is safe.
        width = len(self.headers)
        self.raw = [
            row + [None] * (width - len(row)) if len(row) < width else row
            for row in self.raw
        ]

        last_idx = None  # 0-based index of last row with a value in column A
        for i in range(len(self.raw) - 1, header_idx, -1):
            if self.raw[i] and self.raw[i][0] not in (None, ""):
                last_idx = i
                break
        if last_idx is None:
            raise ValueError(f"{label} file: no data rows found under the header.")

        self.header_row = header_idx + 1  # 1-based Excel rows
        self.first_data_row = header_idx + 2
        self.last_data_row = last_idx + 1

    def col_letter(self, header: str) -> str:
        try:
            return get_column_letter(self.headers.index(header) + 1)
        except ValueError:
            raise ValueError(f"{self.label} file: '{header}' column not found.")


def _excel_num(v) -> float:
    """Coerce a cell value to a number, recovering Excel serials from cells
    that are stored as numbers but formatted as dates/times."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, datetime):
        return (v - datetime(1899, 12, 30)).total_seconds() / 86400.0
    if isinstance(v, date):
        return float((v - date(1899, 12, 30)).days)
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
    return float(v)


def _pivot(report: _Report, code_header: str, value_header: str):
    """Group the report by code and sum the value column, Excel-pivot style.

    Returns (rows, blank_sum, grand_total): rows is [(code, sum)] sorted
    ascending; blank_sum is the (blank) bucket (rows with a value but no
    code) or None when absent, matching what an Excel pivot would show.
    """
    try:
        ci = report.headers.index(code_header)
        vi = report.headers.index(value_header)
    except ValueError as e:
        raise ValueError(f"{report.label} file: column not found ({e}).")

    sums: dict[str, float] = {}
    blank_sum = None
    for row in report.raw[report.first_data_row - 1 :]:
        if not row:
            continue
        code = row[ci] if len(row) > ci else None
        val = row[vi] if len(row) > vi else None
        if code not in (None, ""):
            sums[str(code)] = sums.get(str(code), 0.0) + _excel_num(val)
        elif val is not None and val != "":
            blank_sum = (blank_sum or 0.0) + _excel_num(val)

    rows = sorted(sums.items())
    grand_total = sum(sums.values()) + (blank_sum or 0.0)
    return rows, blank_sum, grand_total


def build_bum_workbook(
    main_bytes: bytes,
    auh_bytes: bytes,
    renewals_bytes: bytes,
    insurance_bytes: bytes,
    pdc_bytes: bytes,
    backlog_bytes: bytes,
):
    """Build the enriched workbook. Returns (xlsx_bytes, meta dict)."""
    as_of, headers, data = _read_main(main_bytes)
    auh = _Report(auh_bytes, "Cust Code", "AUH")
    auh_state = auh.col_letter("Addr State Code")
    renewals = _Report(renewals_bytes, "Invoice number", "Renewals")
    insurance = _Report(insurance_bytes, "Customer Code", "Insurance")
    ins_limit = insurance.col_letter("Insurance Limit")
    pdc_pivot = _pivot(
        _Report(pdc_bytes, "Division", "PDC"), "Sub Account", "LC Amount"
    )
    backlog_pivot = _pivot(
        _Report(backlog_bytes, "Order", "Backlog"),
        "Customer Code",
        "Pending Val (Lc)",
    )

    bum_fixed = _read_csv("bum_fixed.csv")
    region = _read_csv("region.csv")
    gsi = _read_csv("gsi.csv")
    se_africa = _read_csv("se_africa.csv")

    # Column letters of the input fields used in formulas.
    L = {
        name: get_column_letter(headers.index(name) + 1)
        for name in _REQUIRED_HEADERS
    }
    # Column letters of the new columns.
    N = {
        name: get_column_letter(len(headers) + i + 1)
        for i, name in enumerate(NEW_HEADERS)
    }

    eom = date(
        as_of.year, as_of.month, calendar.monthrange(as_of.year, as_of.month)[1]
    )
    eom_ref = f"DATE({eom.year},{eom.month},{eom.day})"

    first_data_row = 3
    last_data_row = first_data_row + len(data) - 1

    def row_formulas(n: int) -> dict[str, str]:
        age, val = N["Invoice Age"], N["Invoice Value"]
        a15, a30, a60 = N["3. Aging 1 to 15"], N["4. Aging 16 to 30"], N["5. Aging 31 to 60"]
        a90, a120, a121 = N["6. Aging 61 to 90"], N["7. Aging 91 to 120"], N["8. Aging >=121"]
        onacc, notdue = N["1. On Account"], N["2. Not Due"]
        code, creg, brand = L["Cust Code"], L["Cust Region"], L["Brand"]
        due, ar = L["Document Due Date"], L["Ar Balance"]
        return {
            "Invoice Age": f"={L['Over Due Days']}{n}",
            "Invoice Value": f"=IF({ar}{n}>0,{ar}{n},0)",
            "3. Aging 1 to 15": (
                f"=IF(${age}{n}>=0,${val}{n},0)"
                f"-{a30}{n}-{a60}{n}-{a90}{n}-{a120}{n}-{a121}{n}"
            ),
            "4. Aging 16 to 30": (
                f"=IF(${age}{n}>=16,${val}{n},0)"
                f"-{a60}{n}-{a90}{n}-{a120}{n}-{a121}{n}"
            ),
            "5. Aging 31 to 60": (
                f"=IF(${age}{n}>=31,${val}{n},0)-{a90}{n}-{a120}{n}-{a121}{n}"
            ),
            "6. Aging 61 to 90": (
                f"=IF(${age}{n}>=61,${val}{n},0)-{a120}{n}-{a121}{n}"
            ),
            "7. Aging 91 to 120": f"=IF(${age}{n}>=91,${val}{n},0)-{a121}{n}",
            "8. Aging >=121": f"=IF(${age}{n}>=121,${val}{n},0)",
            "Additional due End of month": (
                f"=IF(AND({due}{n}>$A$1,{due}{n}<={eom_ref}),{notdue}{n},0)"
            ),
            "9. AR Balance": (
                f"=SUM({a15}{n}:{a121}{n})+SUM({onacc}{n}:{notdue}{n})"
            ),
            "1. On Account": f"={L['On Account']}{n}",
            "2. Not Due": f"={L['Not Due Amount']}{n}",
            "AUH": (
                f"=XLOOKUP({code}{n},"
                f"Sheet3!$A${auh.first_data_row}:$A${auh.last_data_row},"
                f"Sheet3!${auh_state}${auh.first_data_row}"
                f":${auh_state}${auh.last_data_row},"
                f'"NOT AUH")'
            ),
            "BUM": (
                f"=XLOOKUP({brand}{n},"
                f"'BUM fixed'!$A$2:$A${len(bum_fixed)},"
                f"'BUM fixed'!$B$2:$B${len(bum_fixed)},\"\")"
            ),
            "Region": (
                f'=IF(LEFT(TRIM({code}{n}),2)="CK","KSA",'
                f"XLOOKUP({creg}{n},"
                f"Region!$A$2:$A${len(region)},Region!$B$2:$B${len(region)}))"
            ),
            "Shelly": (
                f"=XLOOKUP({code}{n},"
                f"GSI!$A$2:$A${len(gsi)},GSI!$C$2:$C${len(gsi)},\"Not GSI\")"
            ),
            "Renewals": (
                f"=IFERROR(VLOOKUP(${L['Document Number']}{n},"
                f"'Renewal invoices'!$A:$G,7,0),\"Not Renewals\")"
            ),
            "SE Africa": (
                f"=XLOOKUP({creg}{n},"
                f"'SE Africa'!$A$2:$A${len(se_africa)},"
                f"'SE Africa'!$B$2:$B${len(se_africa)},\"NOT SE AFRICA\")"
            ),
            "Insurance": (
                f"=XLOOKUP({code}{n},"
                f"Insurance!$A${insurance.first_data_row}"
                f":$A${insurance.last_data_row},"
                f"Insurance!${ins_limit}${insurance.first_data_row}"
                f":${ins_limit}${insurance.last_data_row},0)"
            ),
            "Backlog": (
                f"=XLOOKUP({code}{n},"
                f"'Backlog Pivot'!$A$4:$A${3 + len(backlog_pivot[0])},"
                f"'Backlog Pivot'!$B$4:$B${3 + len(backlog_pivot[0])},0)"
            ),
            "PDC": (
                f"=XLOOKUP({code}{n},"
                f"'PDC Pivot'!$A$4:$A${3 + len(pdc_pivot[0])},"
                f"'PDC Pivot'!$B$4:$B${3 + len(pdc_pivot[0])},0)"
            ),
        }

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(
        buf,
        {"constant_memory": True, "default_date_format": _FMT_DATE},
    )
    fmt_date = wb.add_format({"num_format": _FMT_DATE})
    fmt_hdr = wb.add_format({"bold": True})
    fmt_hdr_new = wb.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#002060"}
    )
    fmt_usd = wb.add_format({"num_format": _FMT_USD})
    fmt_num = wb.add_format({"num_format": _FMT_NUM})
    fmt_num2 = wb.add_format({"num_format": _FMT_NUM2})
    fmt_plain = wb.add_format({"num_format": _FMT_PLAIN})

    # ── Sheet1: main dump + formula columns ────────────────────────────────
    ws = wb.add_worksheet("Sheet1")

    def col_idx(letter: str) -> int:
        return openpyxl.utils.column_index_from_string(letter) - 1

    # Column formats/widths (must precede rows in constant_memory mode).
    for name in NEW_HEADERS[1:12]:  # Invoice Value .. 2. Not Due
        ws.set_column(f"{N[name]}:{N[name]}", 14, fmt_usd)
    ws.set_column(f"{N['Invoice Age']}:{N['Invoice Age']}", 10)
    for name in ("AUH", "BUM", "Region", "Shelly", "Renewals", "SE Africa"):
        ws.set_column(f"{N[name]}:{N[name]}", 12)
    for name in ("Insurance", "Backlog", "PDC"):
        ws.set_column(f"{N[name]}:{N[name]}", 10, fmt_plain)

    # Row 1: as-of date + control totals.
    ws.write_datetime(0, 0, as_of, fmt_date)
    add_l, bal_l = N["Additional due End of month"], N["9. AR Balance"]
    ws.write_formula(
        0,
        col_idx(add_l),
        f"=SUBTOTAL(9,{add_l}{first_data_row}:{add_l}{last_data_row})",
        fmt_num,
    )
    ws.write_formula(
        0,
        col_idx(bal_l),
        f"=SUM({bal_l}{first_data_row}:{bal_l}{last_data_row})"
        f"-SUM({L['Ar Balance']}{first_data_row}:{L['Ar Balance']}{last_data_row})",
        fmt_num2,
    )

    # Row 2: headers.
    for c, h in enumerate(headers):
        ws.write(1, c, h, fmt_hdr)
    for i, h in enumerate(NEW_HEADERS):
        ws.write(1, len(headers) + i, h, fmt_hdr_new)

    # Data rows.
    for r, row in enumerate(data):
        excel_row = first_data_row + r
        for c, v in enumerate(row):
            if v is None:
                continue
            ws.write(excel_row - 1, c, v)
        for name, formula in row_formulas(excel_row).items():
            ws.write_formula(excel_row - 1, col_idx(N[name]), formula)

    # ── Report copies (lookup sources) ─────────────────────────────────────
    # constant_memory flushes each row once written, so any extra cell must
    # be written in the same row pass as the raw copy.
    def write_report(sheet_name, report, widths, extra=None):
        wsx = wb.add_worksheet(sheet_name)
        for rng, w in widths.items():
            wsx.set_column(rng, w)
        for r, row in enumerate(report.raw):
            for c, v in enumerate(row):
                if v is None:
                    continue
                wsx.write(r, c, v)
            if extra:
                for c, v, fmt in extra(r + 1, row):
                    wsx.write(r, c, v, fmt)
        return wsx

    write_report("Sheet3", auh, {"A:A": 12, "B:B": 42, "C:D": 15})

    # Helper column G: the Renewals VLOOKUP returns column 7 of A:G, which
    # the raw report does not have. Skip if the report ever grows a column G.
    def renewals_helper(excel_row, row):
        if len(renewals.headers) >= 7 and renewals.headers[6]:
            return
        if excel_row == renewals.header_row:
            yield 6, "Renewals", fmt_hdr
        elif (
            renewals.first_data_row <= excel_row <= renewals.last_data_row
            and row
            and row[0] not in (None, "")
        ):
            yield 6, "Renewals", None

    write_report(
        "Renewal invoices",
        renewals,
        {"A:B": 15, "C:C": 42, "D:F": 14, "G:G": 12},
        extra=renewals_helper,
    )

    write_report("Insurance", insurance, {"A:A": 12, "B:B": 42, "C:L": 14})

    # ── Computed pivot sheets (Excel-pivot layout, data from row 4) ────────
    for sheet_name, value_header, (rows, blank_sum, grand_total) in (
        ("PDC Pivot", "Sum of LC Amount", pdc_pivot),
        ("Backlog Pivot", "Sum of Pending Val (Lc)", backlog_pivot),
    ):
        wsx = wb.add_worksheet(sheet_name)
        wsx.set_column("A:A", 14)
        wsx.set_column("B:B", 22)
        wsx.write(2, 0, "Row Labels", fmt_hdr)
        wsx.write(2, 1, value_header, fmt_hdr)
        r = 3
        for code, total in rows:
            wsx.write(r, 0, code)
            wsx.write(r, 1, total)
            r += 1
        if blank_sum is not None:
            wsx.write(r, 0, "(blank)")
            wsx.write(r, 1, blank_sum)
            r += 1
        wsx.write(r, 0, "Grand Total", fmt_hdr)
        wsx.write(r, 1, grand_total, fmt_hdr)

    # ── Fixed lookup sheets ────────────────────────────────────────────────
    for sheet_name, rows, widths in (
        ("BUM fixed", bum_fixed, (26, 18)),
        ("Region", region, (32, 10)),
        ("GSI", gsi, (12, 62, 8)),
        ("SE Africa", se_africa, (30, 12)),
    ):
        wsx = wb.add_worksheet(sheet_name)
        for c, w in enumerate(widths):
            wsx.set_column(c, c, w)
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                wsx.write(r, c, v, fmt_hdr if r == 0 else None)

    wb.close()

    meta = {
        "as_of": as_of,
        "eom": eom,
        "rows": len(data),
        "auh_rows": auh.last_data_row - auh.first_data_row + 1,
        "renewal_rows": renewals.last_data_row - renewals.first_data_row + 1,
        "insurance_rows": insurance.last_data_row - insurance.first_data_row + 1,
        "pdc_customers": len(pdc_pivot[0]),
        "pdc_total": pdc_pivot[2],
        "backlog_customers": len(backlog_pivot[0]),
        "backlog_total": backlog_pivot[2],
    }
    return buf.getvalue(), meta
